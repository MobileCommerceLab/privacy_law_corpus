### WORD STATISTICS 

import os
from openpyxl import Workbook 
import openpyxl 
import statistics

workbook = Workbook()
workbook.save(filename="GPI-word-count-native-english.xlsx")

wb = openpyxl.Workbook() 
sheet = wb.active

c1 = sheet.cell(row = 1, column = 1) 
c1.value = "S.no"
c2 = sheet.cell(row = 1, column = 2) 
c2.value = "File name"
c3 = sheet.cell(row = 1, column = 3) 
c3.value = "Word Count"

## Storing the name of the Directory to traverse and identify files##
p1 = "F:/Word-statistics/native_english_text_files/" #directory 1       ###CHANGE PATH BEFORE RUNNING THE CODE
p2 = "F:/Word-statistics/english_translated_text_files/" #directory 2   ###CHANGE PATH BEFORE RUNNING THE CODE


## Getting names of all files present in the directories ##
dir_list_1=os.listdir(p1)
len1 = len(dir_list_1) 

dir_list_2=os.listdir(p2)
len2 = len(dir_list_2)

storage1 = []  ## Create empty list to store the values 

####WORD COUNT EXCEL FILE FOR NATIVE ENGLISH DOCUMENTS####
for i in range(0,len1):
    path = p1 + dir_list_1[i]
    #path2= "english_translated_text_files/" + dir_list_2
    
    count = 0
    with open(path, "rb") as file:
        data = file.read()
        lines = data.split()
        count += len(lines)
        temp = str(i+1) 
    storage1.append(count)
    #output=[dir_list_1[i], count]
    #print(dir_list_1[i], count)
    c3 = sheet['A'+str(i+2)]
    c3.value = i+1 
    c4 = sheet['B'+str(i+2)]
    c4.value = dir_list_1[i]
    c5 = sheet['C'+str(i+2)]
    c5.value = count

workbook.save(filename="GPI-word-count-native-english.xlsx")
wb.save("GPI-word-count-native-english.xlsx")


####WORD COUNT EXCEL FILE FOR TRANSLATED DOCUMENTS####
workbook.save(filename="GPI-word-count-translated.xlsx")

wb = openpyxl.Workbook() 
sheet = wb.active
c1 = sheet.cell(row = 1, column = 1) 
c1.value = "S.no"
c2 = sheet.cell(row = 1, column = 2) 
c2.value = "File name"
c3 = sheet.cell(row = 1, column = 3) 
c3.value = "Word Count"



for i in range(0,len2):
    path = p2 + dir_list_2[i]
    count = 0
    with open(path, "rb") as file:
        data = file.read()
        lines = data.split()
        count += len(lines)
        temp = str(i+1)
    storage1.append(count) 
    #output=[dir_list_1[i], count]
    #print(dir_list_1[i], count)
    c3 = sheet['A'+str(i+2)]
    c3.value = i+1 
    c4 = sheet['B'+str(i+2)]
    c4.value = dir_list_2[i]
    c5 = sheet['C'+str(i+2)]
    c5.value = count
workbook.save(filename="GPI-word-count-translated.xlsx")
wb.save("GPI-word-count-translated.xlsx")


# Count Computations
#print(len(storage1))
sum = 0
storage1.sort()
del storage1[0]
min_val = min(storage1)
max_val = max(storage1)
median_val = statistics.median(storage1)
mean_val = statistics.mean(storage1) 

for i in range(len(storage1)):
    sum = sum + storage1[i]

print("total words:", sum)
print("min words in a file:", min_val)
print("max words in a file:", max_val)
print("median is:", median_val)
print("mean is:", mean_val)

with open("stats.txt","w") as file:
    file.write("total words:" + str(sum))
    file.write("\nmin words in a file:" + str(min_val))
    file.write("\nmax words in a file:" + str(max_val))
    file.write("\nmedian is:" + str(median_val))
    file.write("\nmean is:" + str(mean_val))
